import httpx
import orjson
from openpyxl import load_workbook


class ExcelIngestionService:

    async def stream_and_push(self, ingestion_id: str, request):
        chunk = []
        chunk_number = 0
        total_records = 0

        wb = load_workbook(
            filename=request.file_path,
            read_only=True,
            data_only=True
        )

        sheet = wb.active
        rows = sheet.iter_rows(values_only=True)

        header_row = next(rows, None)
        if not header_row:
            wb.close()
            return

        headers = [
            str(col).strip() if col is not None else f"column_{i}"
            for i, col in enumerate(header_row)
        ]

        async with httpx.AsyncClient(timeout=60) as client:
            for row in rows:
                record = {
                    headers[i]: row[i]
                    for i in range(len(headers))
                }

                chunk.append(record)
                total_records += 1

                if len(chunk) >= request.chunk_size_by_records:
                    await self._send_chunk(
                        client,
                        request.callback_url,
                        ingestion_id,
                        chunk_number,
                        chunk,
                        False
                    )
                    chunk_number += 1
                    chunk.clear()

            if chunk:
                await self._send_chunk(
                    client,
                    request.callback_url,
                    ingestion_id,
                    chunk_number,
                    chunk,
                    True
                )

            await client.post(
                request.callback_url,
                json={
                    "ingestion_id": ingestion_id,
                    "status": "COMPLETED",
                    "total_records": total_records
                }
            )

        wb.close()

    async def _send_chunk(
        self,
        client,
        url,
        ingestion_id,
        chunk_number,
        records,
        is_last
    ):
        payload = {
            "ingestion_id": ingestion_id,
            "chunk_number": chunk_number,
            "records": records,
            "is_last": is_last
        }

        resp = await client.post(
            url,
            content=orjson.dumps(payload),
            headers={"Content-Type": "application/json"}
        )
        resp.raise_for_status()
