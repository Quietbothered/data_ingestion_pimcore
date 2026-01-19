import httpx
import orjson
from openpyxl import load_workbook

from app.utils.logger import LoggerFactory
from app.services.data_integrity_manager import ChunkIntegrityManager
from app.utils.logger_info_messages import ExcelInfoMessages
from app.utils.error_messages import ExcelErrorMessages

# Import the state store utility
from app.services.ingestion_state_store import IngestionStateStore

info_logger = LoggerFactory.get_info_logger()
error_logger = LoggerFactory.get_error_logger()
debug_logger = LoggerFactory.get_debug_logger()


class ExcelIngestionService:

    def __init__(self):
        self.state_store = IngestionStateStore()
        self.total_records = 0

    async def stream_and_push(self, ingestion_id: str, request):
        # Recover state from DB
        last_chunk = self.state_store.get_last_chunk(ingestion_id)  # last ACKed chunk num (or -1)
        # next chunk number to attempt to send
        chunk_number = last_chunk + 1

        # total_records is authoritative: number of records already ACKed (not raw rows)
        self.total_records = self.state_store.get_total_records(ingestion_id) or 0
        records_to_skip = int(self.total_records)  # number of non-empty records already processed

        chunk = []
        info_logger.info(ExcelInfoMessages.STREAM_START.value.format(ingestion_id=ingestion_id))
        info_logger.info(ExcelInfoMessages.WORKBOOK_LOAD_START.value)

        wb = load_workbook(filename=request.file_path, read_only=True, data_only=True)
        info_logger.info(ExcelInfoMessages.WORKBOOK_LOADED.value)

        sheet = wb.active
        rows = sheet.iter_rows(values_only=True)

        # header
        header_row = next(rows, None)
        debug_logger.debug(f"Header row detected | header_row={header_row}")

        if not header_row:
            error_logger.error(ExcelErrorMessages.EMPTY_HEADER.value.format(ingestion_id=ingestion_id))
            wb.close()
            return

        headers = [str(col).strip() if col is not None else f"column_{i}" for i, col in enumerate(header_row)]
        debug_logger.debug(f"Headers parsed | headers={headers}")

        # We will skip 'records_to_skip' non-empty rows (not raw rows), because earlier runs may have skipped empties.
        skipped_records = 0

        async with httpx.AsyncClient(timeout=60) as client:
            for row in rows:
                # ignore completely empty rows (they don't count toward processed-records)
                if not any(row):
                    continue

                # If we haven't yet skipped up to the saved count, keep skipping
                if skipped_records < records_to_skip:
                    skipped_records += 1
                    # keep chunk_number as-is (we haven't produced any new chunk here)
                    continue

                # This is a new record to process
                record = {headers[i]: row[i] if i < len(row) else None for i in range(len(headers))}
                chunk.append(record)
                self.total_records += 1  # increment only for newly processed record

                # If we have a configured chunk-size-by-records, flush when reached
                if request.chunk_size_by_records and len(chunk) >= request.chunk_size_by_records:
                    # Only send if this chunk hasn't been ACKed yet
                    if chunk_number > last_chunk:
                        debug_logger.debug(
                            f"Chunk processing | ingestion_id={ingestion_id} | "
                            f"chunk_number={chunk_number} | size={len(chunk)} | action=SENDING"
                        )
                        await self._send_chunk(
                            client,
                            request.callback_url,
                            ingestion_id,
                            chunk_number,
                            chunk,
                            False
                        )
                    else:
                        debug_logger.debug(
                            f"Chunk skipping | ingestion_id={ingestion_id} | "
                            f"chunk_number={chunk_number} | action=SKIPPED (Already ACKed)"
                        )

                    chunk_number += 1
                    chunk.clear()

            # Final chunk (if any)
            if chunk:
                if chunk_number > last_chunk:
                    debug_logger.debug(
                        f"Final chunk created | ingestion_id={ingestion_id} | "
                        f"chunk_number={chunk_number} | size={len(chunk)}"
                    )
                    await self._send_chunk(client, request.callback_url, ingestion_id, chunk_number, chunk, True)
                else:
                    debug_logger.debug(
                        f"Final chunk skipping | ingestion_id={ingestion_id} | "
                        f"chunk_number={chunk_number} | action=SKIPPED (Already ACKed)"
                    )

            # Final completion callback
            info_logger.info(ExcelInfoMessages.INGESTION_COMPLETED.value.format(total_records=self.total_records))

            resp = await client.post(
                request.callback_url,
                json={
                    "ingestion_id": ingestion_id,
                    "status": "COMPLETED",
                    "chunk_number": chunk_number,
                    "total_records": self.total_records,
                }
            )

            ack_response = resp.json()
            ack = ack_response.get("ack")
            if ack:
                self.state_store.mark_completed(ingestion_id)

        wb.close()

    async def _send_chunk(self, client, url, ingestion_id, chunk_number, records, is_last):
        checksum = ChunkIntegrityManager.compute_checksum(records)
        chunk_id = ChunkIntegrityManager.build_chunk_id(ingestion_id, chunk_number)

        payload = {
            "ingestion_id": ingestion_id,
            "chunk_number": chunk_number,
            "chunk_id": chunk_id,
            "checksum": checksum,
            "records": records,
            "is_last": is_last,
        }

        for attempt in range(3):
            try:
                debug_logger.debug(
                    f"Sending chunk | chunk_number={chunk_number} | attempt={attempt + 1} | records={len(records)}"
                )
                resp = await client.post(url, content=orjson.dumps(payload), headers={"Content-Type": "application/json"})
                ack_response = resp.json()
                debug_logger.debug(f"Pimcore callback response | response={ack_response}")
                ack = ack_response.get("ack")

                if ack is not True:
                    error = ack_response.get("error")
                    error_logger.error(
                        ExcelErrorMessages.CHUNK_REJECTED.value.format(chunk_number=chunk_number, reason=error)
                    )
                    raise Exception(f"Chunk {chunk_number} rejected: {error}")

                # Persist progress only after successful ACK
                self.state_store.update_chunk(ingestion_id, chunk_number, self.total_records)
                return

            except Exception as e:
                error_logger.error(
                    ExcelErrorMessages.CHUNK_PUSH_FAILED.value.format(
                        chunk_number=chunk_number, attempt=attempt + 1, error=str(e)
                    ),
                    exc_info=True,
                )
                if attempt == 2:
                    raise
